# @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@    Python 3 Script - ASTRO Assignment      @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

# import packages -----------------------------
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl import utils
from datetime import datetime
from datetime import date
import os
import re
# ---------------------------------------------
print()
print(f'((((((((((((((( Extract Address details Process on Datetime: {datetime.now()} )))))))))))))))')
print()
###############################################################################################################
# +++++++++++++++++++++++++++++++++  Folder & File Settings ++++++++++++++++++++++++++++++++++++++++
#Mention Source folder full path- all single backslash need to replace with doubble back slash as below.
#source_folder = f"D:\\Shiva\\Profile\\Calls\\Astro\\astrotechnicaltestrpa"
source_folder = f"C:\\Users\\corp.rpa\\Desktop\\Try"

# Geting Source file path-----------------------------------------
if os.path.exists(source_folder):
    source = f'{source_folder}\\{os.listdir(source_folder)[0]}' # Fetching file
    report = f"{source_folder}\\Result.xlsx"  # Report file gets created - in the same Source folder.
    print(f'Source file = {source}')
    print(f'Result file = {report}')
else:
    print(f'Error: source_folder not found. Make sure folder path is valid')
# ++++++++++++++++++++++++++++  Folder & File Settings Ends Here ++++++++++++++++++++++++++++++++++++


# --------------------------- Open Excel Sheets and Columns Settings ------------------------------------
s = openpyxl.load_workbook(source)
sSheet =s['Sheet1']
s.save(report) # Saving file as

r = openpyxl.load_workbook(report)
rSheet =r['Sheet1']
r.create_sheet('Sheet2') # Creating new Sheet
rSheet2 =r['Sheet2']    # Used to store sliced address keywords

# From the Source file, Getting Column numbers
s_add_col = 2
s_zip_col = 4

s_rows = sSheet.max_row # Getting total rows count


# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!  FUNCTIONS !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
# Function extracts area Zip codes having 5 digits. Used regex.
def zipcodes(address):
    # print (f'Address= {address}\n')
    #import re
    # Using try catch block to handle errors, may occur if 5 digits pattern is not available in the address.
    try:
        zips = re.findall('[\d]{5}', address)  # Fetching digits having 5 in count
        #print(f'Zips= {zips[0]}')
        return zips[0]
    except:
        #print(f'Zips= NOT FOUND')
        return 'NOT FOUND'

# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

#------ Writting Report headder --------------------
rSheet.cell(1, 1).value = 'S.No'
rSheet.cell(1, 2).value = 'Full Address'
rSheet.cell(1, 3).value = 'Time Network'
rSheet.cell(1, 4).value = 'Maxis Network'
rSheet2.cell(1, 1).value = 'S.No'
rSheet2.cell(1, 2).value = 'Full Address'
rSheet2.cell(1, 3).value = 'Parts Count'
rSheet2.cell(1, 4).value = 'Zip Code'

print("******************************* Processed below records *******************************")
r_rowNo = 2
for i in range(2, s_rows + 1):
    len = 0
    s_address = sSheet.cell(i, s_add_col).value
    print(f'Address {i-1}: {s_address}')
    zip = zipcodes(s_address)
    counter = s_address.count(',')
    #print(f'Counter= {counter}')
    # Using try catch block to handle errors, may occur if there is no comma in address passed.
    try:
        aSplit = s_address.split(',')
    except:
        counter = 0

    #Writing to the report sheet1
    rSheet.cell(i, 1).value = (i-1)
    rSheet.cell(i, s_add_col).value = s_address
    # Writing to the report sheet2
    rSheet2.cell(i, 1).value = (i-1)
    rSheet2.cell(i, s_add_col).value = s_address
    rSheet2.cell(i, s_zip_col-1).value = counter + 1 # count of sliced address parts.
    rSheet2.cell(i, s_zip_col).value = zip
    #r.save(report)

    for j in range(0, counter + 1):
        #print(f'Counter= {counter} : j = {j}: data = {aSplit[j].strip()}')
        rSheet2.cell(i, j+5).value = aSplit[j].strip()  #Writting values to report Sheet2

# $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$ Save Workbooks $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
r.save(report)

print()
print("********************************** Process Completed ***************************************")

print(f'Process Completed on datetime: {datetime.now()}')

# (((((((((((((((((((((   Process Ends Here )))))))))))))))))))))))))
